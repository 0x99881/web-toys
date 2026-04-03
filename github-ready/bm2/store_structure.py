from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .constants import (
    INCOME_META_HEADERS,
    INCOME_META_SHEET,
    INCOME_META_SHEET_ALIASES,
    INCOME_NAME_HEADER,
    LEGACY_WEAR_SHEET,
    META_HEADERS,
    META_SHEET,
    META_SHEET_ALIASES,
    NAME_HEADER,
    SCORE_SHEET,
    TOTAL_HEADER,
    WEAR_META_HEADERS,
    WEAR_META_SHEET,
    WEAR_META_SHEET_ALIASES,
    WEAR_NAME_HEADER,
    WEAR_SHEET,
    WEAR_TOTAL_HEADER,
    WINDOW_SIZE,
)


class StoreStructureMixin:
    def _ensure_score_sheet_structure(self, workbook) -> bool:
        sheet = self._score_sheet(workbook)
        changed = False

        col = 1
        while col <= sheet.max_column:
            if sheet.cell(1, col).value is not None:
                col += 1
                continue
            if any(sheet.cell(row, col).value is not None for row in range(2, sheet.max_row + 1)):
                col += 1
                continue
            sheet.delete_cols(col)
            changed = True

        total_col, name_col, tail_changed = self._ensure_summary_columns(sheet, TOTAL_HEADER, NAME_HEADER)
        changed = changed or tail_changed

        existing_rows = self._build_name_row_map(sheet, name_col)
        row_count_before = len(existing_rows)
        existing_rows = self._ensure_member_rows(
            sheet,
            name_col=name_col,
            total_col=total_col,
            value_columns=[col for _, col in self._d_columns(sheet)],
        )
        changed = changed or len(existing_rows) != row_count_before

        self._recalculate_totals(sheet, total_col)
        self._sort_named_rows(sheet, total_col, name_col)
        recent_numbers = [number for number, _ in self._d_columns(sheet)][-WINDOW_SIZE:]
        self._format_score_sheet(sheet, recent_numbers, total_col)
        return changed

    def _recalculate_wear_totals(self, sheet, total_col: int) -> None:
        name_col = self._find_column(sheet, WEAR_NAME_HEADER)
        if name_col is None:
            return
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row, name_col).value:
                continue
            total = 0.0
            for _, col in self._wear_columns(sheet):
                value = sheet.cell(row, col).value
                numeric = self._round_wear(value or 0)
                sheet.cell(row, col, numeric)
                total += numeric
            sheet.cell(row, total_col, self._round_wear(total))

    def _ensure_wear_sheet_structure(self, workbook) -> bool:
        sheet = self._wear_sheet(workbook)
        changed = False

        legacy_headers = ["日期", "姓名", "每日磨损"]
        current_headers = [sheet.cell(1, idx).value for idx in range(1, 4)]
        if current_headers == legacy_headers:
            self._migrate_legacy_wear_rows(workbook, sheet)
            changed = True

        total_col, name_col, tail_changed = self._ensure_summary_columns(sheet, WEAR_TOTAL_HEADER, WEAR_NAME_HEADER)
        changed = changed or tail_changed

        existing_rows = self._build_name_row_map(sheet, name_col)
        row_count_before = len(existing_rows)
        existing_rows = self._ensure_member_rows(
            sheet,
            name_col=name_col,
            total_col=total_col,
            value_columns=[col for _, col in self._wear_columns(sheet)],
        )
        changed = changed or len(existing_rows) != row_count_before

        self._recalculate_wear_totals(sheet, total_col)
        self._sort_named_rows(sheet, total_col, name_col)
        return changed

    def _ensure_income_sheet_structure(self, workbook) -> bool:
        sheet = self._income_sheet(workbook)
        changed = False

        name_col = self._find_column(sheet, INCOME_NAME_HEADER)
        if name_col is None:
            sheet.cell(1, sheet.max_column + 1, INCOME_NAME_HEADER)
            changed = True
            name_col = self._find_column(sheet, INCOME_NAME_HEADER)

        assert name_col is not None
        if name_col != sheet.max_column:
            name_values = [sheet.cell(row, name_col).value for row in range(1, sheet.max_row + 1)]
            sheet.delete_cols(name_col)
            insert_at = sheet.max_column + 1
            sheet.insert_cols(insert_at, 1)
            for row, value in enumerate(name_values, start=1):
                sheet.cell(row, insert_at, value)
            changed = True
            name_col = insert_at

        existing_rows = self._build_name_row_map(sheet, name_col)
        row_count_before = len(existing_rows)
        existing_rows = self._ensure_member_rows(
            sheet,
            name_col=name_col,
            total_col=None,
            value_columns=[col for _, col in self._income_columns(sheet)],
        )
        changed = changed or len(existing_rows) != row_count_before
        self._sort_rows_by_name(sheet, name_col)
        return changed

    # ?????????????????
    def _ensure_workbook(self) -> None:
        changed = False
        if self.workbook_path.exists():
            workbook = load_workbook(self.workbook_path)
        else:
            workbook = Workbook()
            workbook.active.title = SCORE_SHEET
            changed = True

        _, meta_changed = self._ensure_named_sheet(workbook, META_SHEET, META_SHEET_ALIASES, META_HEADERS, hidden=True)
        changed = meta_changed or changed
        _, wear_meta_changed = self._ensure_named_sheet(workbook, WEAR_META_SHEET, WEAR_META_SHEET_ALIASES, WEAR_META_HEADERS, hidden=True)
        changed = wear_meta_changed or changed
        _, income_meta_changed = self._ensure_named_sheet(workbook, INCOME_META_SHEET, INCOME_META_SHEET_ALIASES, INCOME_META_HEADERS, hidden=True)
        changed = income_meta_changed or changed
        changed = self._ensure_score_sheet_structure(workbook) or changed
        changed = self._ensure_wear_sheet_structure(workbook) or changed
        changed = self._ensure_income_sheet_structure(workbook) or changed
        if changed:
            self._save_workbook(workbook)
        workbook.close()

    def _append_meta(self, workbook, date_text: str, col_name: str) -> None:
        sheet = workbook[META_SHEET]
        sheet.append([date_text, col_name])

    def _append_wear_meta(self, workbook, date_text: str, col_name: str) -> None:
        sheet = workbook[WEAR_META_SHEET]
        sheet.append([date_text, col_name])

    def _append_income_meta(self, workbook, date_text: str, col_name: str) -> None:
        sheet = workbook[INCOME_META_SHEET]
        sheet.append([date_text, col_name])

    def _read_sheet_meta(self, workbook, sheet_name: str) -> list[tuple[str, str]]:
        if sheet_name not in workbook.sheetnames:
            return []
        rows = []
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                rows.append((str(row[0]), str(row[1])))
        return rows

    def _normalize_wear_date_text(self, value: str, year: int) -> str:
        if len(value) == 4 and value.isdigit():
            return f"{year:04d}-{value[:2]}-{value[2:]}"
        return value

    def _migrate_legacy_wear_rows(self, workbook, sheet) -> None:
        legacy_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            legacy_rows.append(
                {
                    "date": str(row[0]),
                    "name": str(row[1]),
                    "wear": float(row[2] or 0),
                }
            )

        sheet.delete_rows(1, sheet.max_row)
        meta_sheet = workbook[WEAR_META_SHEET] if WEAR_META_SHEET in workbook.sheetnames else workbook.create_sheet(WEAR_META_SHEET)
        meta_sheet.delete_rows(1, meta_sheet.max_row)
        meta_sheet.append(WEAR_META_HEADERS)

        date_order: list[str] = []
        member_names: list[str] = []
        data_map: dict[tuple[str, str], float] = {}
        for item in legacy_rows:
            date_text = item["date"]
            member_name = item["name"]
            if date_text not in date_order:
                date_order.append(date_text)
            if member_name not in member_names:
                member_names.append(member_name)
            data_map[(member_name, date_text)] = item["wear"]

        for index, date_text in enumerate(date_order, start=1):
            sheet.cell(1, index, date_text[5:].replace("-", ""))
            meta_sheet.append([date_text, str(index)])

        total_col = len(date_order) + 1
        name_col = total_col + 1
        sheet.cell(1, total_col, WEAR_TOTAL_HEADER)
        sheet.cell(1, name_col, WEAR_NAME_HEADER)
        for member_name in member_names:
            row = sheet.max_row + 1
            running_total = 0.0
            for index, date_text in enumerate(date_order, start=1):
                wear = self._round_wear(data_map.get((member_name, date_text), 0.0))
                sheet.cell(row, index, wear)
                running_total += wear
            sheet.cell(row, total_col, self._round_wear(running_total))
            sheet.cell(row, name_col, member_name)

    def _recalculate_totals(self, sheet, total_col: int) -> None:
        name_col = self._find_column(sheet, NAME_HEADER)
        if name_col is None:
            return
        d_cols = self._d_columns(sheet)
        recent_numbers = {number for number, _ in d_cols[-WINDOW_SIZE:]}
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row, name_col).value:
                continue
            total = 0
            for number, d_col in d_cols:
                value = sheet.cell(row, d_col).value
                numeric = int(value or 0)
                sheet.cell(row, d_col, numeric)
                if number in recent_numbers:
                    total += numeric
            sheet.cell(row, total_col, total)

    def _format_score_sheet(self, sheet, recent_numbers: list[int], total_col: int) -> None:
        old_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        light = (204, 255, 204)
        dark = (0, 170, 0)
        recent_set = set(recent_numbers)
        for number, d_col in self._d_columns(sheet):
            is_recent = number in recent_set
            sheet.cell(1, d_col, f"D{number}" if is_recent else f"旧-D{number}")
            fill = PatternFill(fill_type=None) if is_recent else old_fill
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, d_col).fill = fill

        totals = []
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, total_col).value
            if isinstance(value, (int, float)):
                totals.append(float(value))
        max_score = max(totals) if totals else 0.0
        min_score = min(totals) if totals else 0.0
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, total_col).value
            score = float(value) if isinstance(value, (int, float)) else 0.0
            if max_score == min_score:
                color = "99FF99"
            else:
                ratio = (score - min_score) / (max_score - min_score)
                red = int(light[0] + (dark[0] - light[0]) * ratio)
                green = int(light[1] + (dark[1] - light[1]) * ratio)
                blue = int(light[2] + (dark[2] - light[2]) * ratio)
                color = f"{red:02X}{green:02X}{blue:02X}"
            sheet.cell(row, total_col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

